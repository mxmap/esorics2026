"""Tests for the XLSX export module."""

from __future__ import annotations

import json
from pathlib import Path

import openpyxl

from mail_municipalities.export import (
    export_xlsx,
    flatten_provider,
    flatten_security,
    load_json,
)

# ---------------------------------------------------------------------------
# Synthetic test data
# ---------------------------------------------------------------------------

_PROVIDER_ENTRY_FULL = {
    "code": "1",
    "name": "Zurich",
    "region": "ZH",
    "domain": "zurich.ch",
    "provider": "microsoft",
    "category": "us-cloud",
    "classification_confidence": 95.0,
    "classification_signals": [
        {"kind": "mx", "detail": "outlook"},
        {"kind": "spf", "detail": "spf.protection.outlook.com"},
    ],
    "sources_detail": {"web": True, "wikidata": True},
    "resolve_flags": ["scraped", "validated"],
    "mx": ["mail.protection.outlook.com"],
    "spf": "v=spf1 include:spf.protection.outlook.com -all",
    "gateway": "seppmail",
}

_PROVIDER_ENTRY_MINIMAL = {
    "code": "2",
    "name": "Bern",
    "region": "BE",
    "domain": "bern.ch",
    "provider": "unknown",
    "category": "unknown",
    "classification_confidence": 50.0,
    "classification_signals": [],
    "mx": [],
    "spf": "",
}

_SECURITY_ENTRY_FULL = {
    "code": "1",
    "domain": "zurich.ch",
    "dane": {"supported": True, "partial": False},
    "dss": {
        "has_spf": True,
        "has_good_spf": True,
        "has_dmarc": True,
        "has_good_dmarc": False,
        "has_dkim": True,
    },
    "scan_valid": True,
}

_SECURITY_ENTRY_NONE = {
    "code": "2",
    "domain": "bern.ch",
    "dane": None,
    "dss": None,
    "scan_valid": False,
}

_PROVIDER_DATA = {
    "generated": "2026-01-01T00:00:00Z",
    "total": 2,
    "municipalities": [_PROVIDER_ENTRY_FULL, _PROVIDER_ENTRY_MINIMAL],
}

_SECURITY_DATA = {
    "generated": "2026-01-01T00:00:00Z",
    "total": 2,
    "municipalities": [_SECURITY_ENTRY_FULL, _SECURITY_ENTRY_NONE],
}


def _write_fixtures(tmp_path: Path) -> Path:
    """Write provider + security JSON for all 3 countries, return output_dir."""
    (tmp_path / "providers").mkdir()
    (tmp_path / "security").mkdir()
    for cc in ("ch", "de", "at"):
        (tmp_path / "providers" / f"providers_{cc}.json").write_text(json.dumps(_PROVIDER_DATA), encoding="utf-8")
        (tmp_path / "security" / f"security_{cc}.json").write_text(json.dumps(_SECURITY_DATA), encoding="utf-8")
    return tmp_path


# ---------------------------------------------------------------------------
# load_json
# ---------------------------------------------------------------------------


class TestLoadJson:
    def test_reads_json(self, tmp_path: Path) -> None:
        data = {"key": [1, 2, 3]}
        p = tmp_path / "test.json"
        p.write_text(json.dumps(data), encoding="utf-8")
        assert load_json(p) == data


# ---------------------------------------------------------------------------
# flatten_provider
# ---------------------------------------------------------------------------


class TestFlattenProvider:
    def test_full_fields(self) -> None:
        result = flatten_provider(_PROVIDER_ENTRY_FULL, "CH")
        assert result["country"] == "CH"
        assert result["code"] == "1"
        assert result["name"] == "Zurich"
        assert result["region"] == "ZH"
        assert result["domain"] == "zurich.ch"
        assert result["provider"] == "microsoft"
        assert result["category"] == "us-cloud"
        assert result["confidence"] == 95.0
        assert result["gateway"] == "seppmail"
        assert result["mx"] == "mail.protection.outlook.com"
        assert "spf.protection.outlook.com" in result["spf"]
        assert "mx: outlook" in result["classification_signals"]
        assert "spf: spf.protection.outlook.com" in result["classification_signals"]
        assert "web" in result["sources"]
        assert "wikidata" in result["sources"]
        assert "scraped" in result["resolve_flags"]
        assert "validated" in result["resolve_flags"]

    def test_missing_optional_fields(self) -> None:
        result = flatten_provider(_PROVIDER_ENTRY_MINIMAL, "DE")
        assert result["country"] == "DE"
        assert result["gateway"] == ""
        assert result["mx"] == ""
        assert result["classification_signals"] == ""
        assert result["sources"] == ""
        assert result["resolve_flags"] == ""


# ---------------------------------------------------------------------------
# flatten_security
# ---------------------------------------------------------------------------


class TestFlattenSecurity:
    def test_full_fields(self) -> None:
        result = flatten_security(_SECURITY_ENTRY_FULL)
        assert result["code"] == "1"
        assert result["domain"] == "zurich.ch"
        assert result["scan_valid"] is True
        assert result["dane_supported"] is True
        assert result["dane_partial"] is False
        assert result["has_spf"] is True
        assert result["has_good_spf"] is True
        assert result["has_dmarc"] is True
        assert result["has_good_dmarc"] is False

    def test_none_dane_dss(self) -> None:
        result = flatten_security(_SECURITY_ENTRY_NONE)
        assert result["scan_valid"] is False
        assert result["dane_supported"] is False
        assert result["dane_partial"] is False
        assert result["has_spf"] is False
        assert result["has_good_spf"] is False
        assert result["has_dmarc"] is False
        assert result["has_good_dmarc"] is False

    def test_missing_scan_valid(self) -> None:
        entry = {"code": "3", "domain": "x.ch", "dane": {}, "dss": {}}
        result = flatten_security(entry)
        assert result["scan_valid"] is False


# ---------------------------------------------------------------------------
# export_xlsx
# ---------------------------------------------------------------------------

_EXPECTED_COLUMNS = [
    "country",
    "code",
    "name",
    "region",
    "domain",
    "provider",
    "category",
    "confidence",
    "gateway",
    "mx",
    "spf",
    "scan_valid",
    "dane_supported",
    "dane_partial",
    "has_spf",
    "has_good_spf",
    "has_dmarc",
    "has_good_dmarc",
    "classification_signals",
    "sources",
    "resolve_flags",
]


class TestExportXlsx:
    def test_creates_xlsx(self, tmp_path: Path) -> None:
        output_dir = _write_fixtures(tmp_path)
        result = export_xlsx(output_dir)
        assert result == output_dir / "export.xlsx"
        assert result.exists()

    def test_municipalities_sheet(self, tmp_path: Path) -> None:
        output_dir = _write_fixtures(tmp_path)
        export_xlsx(output_dir)

        wb = openpyxl.load_workbook(output_dir / "export.xlsx")
        ws = wb["Municipalities"]

        # Header row
        headers = [cell.value for cell in ws[1]]
        assert headers == _EXPECTED_COLUMNS

        # 3 countries × 2 munis = 6 data rows
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(data_rows) == 6

        # Sorted by country then code: AT first
        assert data_rows[0][0] == "AT"
        assert data_rows[0][1] == "1"

    def test_statistics_sheet_exists(self, tmp_path: Path) -> None:
        output_dir = _write_fixtures(tmp_path)
        export_xlsx(output_dir)

        wb = openpyxl.load_workbook(output_dir / "export.xlsx")
        assert "Statistics" in wb.sheetnames

        ws2 = wb["Statistics"]
        titles = []
        for row in ws2.iter_rows(min_col=1, max_col=1, values_only=True):
            if row[0] is not None:
                titles.append(row[0])

        assert "Provider Distribution" in titles
        assert "Category Distribution" in titles
        assert "Security" in titles
        assert "Classification Confidence" in titles

    def test_formatting(self, tmp_path: Path) -> None:
        output_dir = _write_fixtures(tmp_path)
        export_xlsx(output_dir)

        wb = openpyxl.load_workbook(output_dir / "export.xlsx")
        ws = wb["Municipalities"]

        # Header is bold
        assert ws["A1"].font.bold is True
        assert ws["B1"].font.bold is True

        # Freeze panes
        assert ws.freeze_panes == "A2"

        # Auto-filter
        assert ws.auto_filter.ref is not None

    def test_column_widths(self, tmp_path: Path) -> None:
        output_dir = _write_fixtures(tmp_path)
        export_xlsx(output_dir)

        wb = openpyxl.load_workbook(output_dir / "export.xlsx")
        ws = wb["Municipalities"]

        assert ws.column_dimensions["A"].width == 8  # country
        assert ws.column_dimensions["C"].width == 30  # name

    def test_statistics_formulas(self, tmp_path: Path) -> None:
        output_dir = _write_fixtures(tmp_path)
        export_xlsx(output_dir)

        wb = openpyxl.load_workbook(output_dir / "export.xlsx")
        ws2 = wb["Statistics"]

        # Collect all formula cells
        formulas = []
        for row in ws2.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas.append(cell.value)

        # Should contain COUNTIF, COUNTIFS, AVERAGE, MEDIAN formulas
        formula_text = " ".join(formulas)
        assert "COUNTIF" in formula_text
        assert "COUNTIFS" in formula_text
        assert "AVERAGE" in formula_text
        assert "MEDIAN" in formula_text

    def test_statistics_freeze_and_widths(self, tmp_path: Path) -> None:
        output_dir = _write_fixtures(tmp_path)
        export_xlsx(output_dir)

        wb = openpyxl.load_workbook(output_dir / "export.xlsx")
        ws2 = wb["Statistics"]

        assert ws2.freeze_panes == "A2"
        assert ws2.column_dimensions["A"].width == 12
