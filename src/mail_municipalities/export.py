import json
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUTPUT_DIR = Path(__file__).parent.parent.parent / "output"


def load_json(path: Path):
    with open(path) as f:
        return json.load(f)


def flatten_provider(m: dict, country: str) -> dict:
    signals = "; ".join(f"{s['kind']}: {s['detail']}" for s in m.get("classification_signals", []))
    sources = "; ".join(m.get("sources_detail", {}).keys())
    flags = "; ".join(m.get("resolve_flags", []))
    return {
        "country": country,
        "code": m["code"],
        "name": m["name"],
        "region": m["region"],
        "domain": m["domain"],
        "provider": m["provider"],
        "category": m["category"],
        "confidence": m["classification_confidence"],
        "gateway": m.get("gateway", ""),
        "mx": "; ".join(m.get("mx", [])),
        "spf": m.get("spf", ""),
        "classification_signals": signals,
        "sources": sources,
        "resolve_flags": flags,
    }


def flatten_security(m: dict) -> dict:
    dane = m.get("dane") or {}
    dss = m.get("dss") or {}
    return {
        "code": m["code"],
        "domain": m["domain"],
        "scan_valid": m.get("scan_valid", False),
        "dane_supported": dane.get("supported", False),
        "dane_partial": dane.get("partial", False),
        "has_spf": dss.get("has_spf", False),
        "has_good_spf": dss.get("has_good_spf", False),
        "has_dmarc": dss.get("has_dmarc", False),
        "has_good_dmarc": dss.get("has_good_dmarc", False),
    }


def export_xlsx(output_dir: Path) -> Path:
    """Build a combined XLSX export from provider and security JSON files."""
    frames = []
    for country in ("ch", "de", "at"):
        provider_data = load_json(output_dir / "providers" / f"providers_{country}.json")
        security_data = load_json(output_dir / "security" / f"security_{country}.json")

        cc = country.upper()
        prov_df = pd.DataFrame([flatten_provider(m, cc) for m in provider_data["municipalities"]])
        sec_df = pd.DataFrame([flatten_security(m) for m in security_data["municipalities"]])

        merged = prov_df.merge(sec_df, on=["code", "domain"], how="left")
        frames.append(merged)

    df = pd.concat(frames, ignore_index=True)
    df.sort_values(["country", "code"], inplace=True, ignore_index=True)

    # Column order
    columns = [
        "country",
        "code",
        "name",
        "region",
        "domain",
        "provider",
        "category",
        "confidence",
        "gateway",
        "mx",
        "spf",
        "scan_valid",
        "dane_supported",
        "dane_partial",
        "has_spf",
        "has_good_spf",
        "has_dmarc",
        "has_good_dmarc",
        "classification_signals",
        "sources",
        "resolve_flags",
    ]
    df = df[columns]

    # Write to XLSX
    out_path = output_dir / "export.xlsx"
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Municipalities")
        ws = writer.sheets["Municipalities"]

        # Bold header
        bold = Font(bold=True)
        for cell in ws[1]:
            cell.font = bold

        # Freeze header row
        ws.freeze_panes = "A2"

        # Auto-filter
        ws.auto_filter.ref = ws.dimensions

        # Column widths
        widths = {
            "country": 8,
            "code": 8,
            "name": 30,
            "region": 25,
            "domain": 30,
            "provider": 12,
            "category": 12,
            "confidence": 12,
            "gateway": 25,
            "mx": 40,
            "spf": 50,
            "scan_valid": 11,
            "dane_supported": 15,
            "dane_partial": 13,
            "has_spf": 9,
            "has_good_spf": 13,
            "has_dmarc": 11,
            "has_good_dmarc": 15,
            "classification_signals": 60,
            "sources": 20,
            "resolve_flags": 15,
        }
        for col_idx, col_name in enumerate(columns, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(col_name, 15)

        # --- Statistics tab (Excel formulas referencing Municipalities sheet) ---
        S = "Municipalities"  # source sheet name
        n = len(df) + 1  # last data row (1-indexed, row 1 is header)
        # Column refs in Municipalities sheet
        COL_COUNTRY = f"'{S}'!A$2:A${n}"
        COL_PROVIDER = f"'{S}'!F$2:F${n}"
        COL_CATEGORY = f"'{S}'!G$2:G${n}"
        COL_CONFIDENCE = f"'{S}'!H$2:H${n}"

        bool_cols = {
            "scan_valid": "L",
            "dane_supported": "M",
            "dane_partial": "N",
            "has_spf": "O",
            "has_good_spf": "P",
            "has_dmarc": "Q",
            "has_good_dmarc": "R",
        }

        wb = writer.book
        ws2 = wb.create_sheet("Statistics")

        title_font = Font(bold=True, size=12)
        header_font = Font(bold=True)
        header_align = Alignment(horizontal="center")
        pct_fmt = "0.0"
        header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

        countries = ["AT", "CH", "DE"]
        providers = ["microsoft", "google", "aws", "domestic", "foreign", "unknown"]
        categories = ["us-cloud", "at-based", "ch-based", "de-based", "foreign", "unknown"]

        def write_header(ws, row, values):
            for c, v in enumerate(values, 1):
                cell = ws.cell(row=row, column=c, value=v)
                cell.font = header_font
                cell.alignment = header_align
                cell.fill = header_fill

        def write_title(ws, row, title):
            cell = ws.cell(row=row, column=1, value=title)
            cell.font = title_font

        row = 1

        # --- Table 1: Provider Distribution ---
        write_title(ws2, row, "Provider Distribution")
        row += 1
        headers = ["Country", "Total"]
        for p in providers:
            headers += [f"{p} (n)", f"{p} (%)"]
        write_header(ws2, row, headers)
        row += 1

        for cc in countries + ["Total"]:
            col = 1
            ws2.cell(row=row, column=col, value=cc).font = Font(bold=(cc == "Total"))
            col += 1
            # Total count
            if cc == "Total":
                ws2.cell(row=row, column=col, value=f"=COUNTA({COL_COUNTRY})")
            else:
                ws2.cell(row=row, column=col, value=f'=COUNTIF({COL_COUNTRY},"{cc}")')
            total_cell = f"${get_column_letter(col)}${row}"
            col += 1
            for p in providers:
                if cc == "Total":
                    ws2.cell(row=row, column=col, value=f'=COUNTIF({COL_PROVIDER},"{p}")')
                else:
                    ws2.cell(row=row, column=col, value=f'=COUNTIFS({COL_COUNTRY},"{cc}",{COL_PROVIDER},"{p}")')
                n_cell = f"{get_column_letter(col)}{row}"
                col += 1
                ws2.cell(row=row, column=col).value = f"={n_cell}/{total_cell}*100"
                ws2.cell(row=row, column=col).number_format = pct_fmt
                col += 1
            row += 1

        row += 1

        # --- Table 2: Category Distribution ---
        write_title(ws2, row, "Category Distribution")
        row += 1
        headers = ["Country", "Total"]
        for cat in categories:
            headers += [f"{cat} (n)", f"{cat} (%)"]
        write_header(ws2, row, headers)
        row += 1

        for cc in countries + ["Total"]:
            col = 1
            ws2.cell(row=row, column=col, value=cc).font = Font(bold=(cc == "Total"))
            col += 1
            if cc == "Total":
                ws2.cell(row=row, column=col, value=f"=COUNTA({COL_COUNTRY})")
            else:
                ws2.cell(row=row, column=col, value=f'=COUNTIF({COL_COUNTRY},"{cc}")')
            total_cell = f"${get_column_letter(col)}${row}"
            col += 1
            for cat in categories:
                if cc == "Total":
                    ws2.cell(row=row, column=col, value=f'=COUNTIF({COL_CATEGORY},"{cat}")')
                else:
                    ws2.cell(row=row, column=col, value=f'=COUNTIFS({COL_COUNTRY},"{cc}",{COL_CATEGORY},"{cat}")')
                n_cell = f"{get_column_letter(col)}{row}"
                col += 1
                ws2.cell(row=row, column=col).value = f"={n_cell}/{total_cell}*100"
                ws2.cell(row=row, column=col).number_format = pct_fmt
                col += 1
            row += 1

        row += 1

        # --- Table 3: Security ---
        write_title(ws2, row, "Security")
        row += 1
        sec_labels = list(bool_cols.keys())
        headers = ["Country", "Total"]
        for label in sec_labels:
            headers += [f"{label} (n)", f"{label} (%)"]
        write_header(ws2, row, headers)
        row += 1

        for cc in countries + ["Total"]:
            col = 1
            ws2.cell(row=row, column=col, value=cc).font = Font(bold=(cc == "Total"))
            col += 1
            if cc == "Total":
                ws2.cell(row=row, column=col, value=f"=COUNTA({COL_COUNTRY})")
            else:
                ws2.cell(row=row, column=col, value=f'=COUNTIF({COL_COUNTRY},"{cc}")')
            total_cell = f"${get_column_letter(col)}${row}"
            col += 1
            for label, src_col in bool_cols.items():
                src_range = f"'{S}'!{src_col}$2:{src_col}${n}"
                if cc == "Total":
                    ws2.cell(row=row, column=col, value=f"=COUNTIF({src_range},TRUE)")
                else:
                    ws2.cell(row=row, column=col, value=f'=COUNTIFS({COL_COUNTRY},"{cc}",{src_range},TRUE)')
                n_cell = f"{get_column_letter(col)}{row}"
                col += 1
                ws2.cell(row=row, column=col).value = f"={n_cell}/{total_cell}*100"
                ws2.cell(row=row, column=col).number_format = pct_fmt
                col += 1
            row += 1

        row += 1

        # --- Table 4: Classification Confidence ---
        write_title(ws2, row, "Classification Confidence")
        row += 1
        write_header(ws2, row, ["Country", "Mean", "Median", "Min", "Max"])
        row += 1

        for cc in countries + ["Total"]:
            col = 1
            ws2.cell(row=row, column=col, value=cc).font = Font(bold=(cc == "Total"))
            col += 1
            if cc == "Total":
                ws2.cell(row=row, column=col, value=f"=AVERAGE({COL_CONFIDENCE})")
                ws2.cell(row=row, column=col).number_format = pct_fmt
                col += 1
                ws2.cell(row=row, column=col, value=f"=MEDIAN({COL_CONFIDENCE})")
                ws2.cell(row=row, column=col).number_format = pct_fmt
                col += 1
                ws2.cell(row=row, column=col, value=f"=MIN({COL_CONFIDENCE})")
                ws2.cell(row=row, column=col).number_format = pct_fmt
                col += 1
                ws2.cell(row=row, column=col, value=f"=MAX({COL_CONFIDENCE})")
                ws2.cell(row=row, column=col).number_format = pct_fmt
            else:
                ws2.cell(row=row, column=col, value=f'=AVERAGEIF({COL_COUNTRY},"{cc}",{COL_CONFIDENCE})')
                ws2.cell(row=row, column=col).number_format = pct_fmt
                col += 1
                # MEDIAN with criteria requires array formula
                ws2.cell(row=row, column=col, value=f'=MEDIAN(IF({COL_COUNTRY}="{cc}",{COL_CONFIDENCE}))')
                ws2.cell(row=row, column=col).number_format = pct_fmt
                col += 1
                ws2.cell(row=row, column=col, value=f'=MIN(IF({COL_COUNTRY}="{cc}",{COL_CONFIDENCE}))')
                ws2.cell(row=row, column=col).number_format = pct_fmt
                col += 1
                ws2.cell(row=row, column=col, value=f'=MAX(IF({COL_COUNTRY}="{cc}",{COL_CONFIDENCE}))')
                ws2.cell(row=row, column=col).number_format = pct_fmt
            row += 1

        # Column widths for statistics sheet
        ws2.column_dimensions["A"].width = 12
        for c in range(2, 20):
            ws2.column_dimensions[get_column_letter(c)].width = 16
        ws2.freeze_panes = "A2"

    print(f"Exported {len(df)} rows to {out_path}")
    return out_path


if __name__ == "__main__":  # pragma: no cover
    export_xlsx(OUTPUT_DIR)
